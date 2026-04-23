from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def exportar_fechamento_excel(mes: str, fechamento, lancamentos, frequencias, funcionarios):
    wb = Workbook()

    ws = wb.active
    wb.remove(ws)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    def estilizar_header(worksheet, row_num=1):
        for cell in worksheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font

    # Resumo Fechamento
    ws_resumo = wb.create_sheet("Resumo Fechamento")
    ws_resumo.append([
        "Funcionário ID",
        "Funcionário",
        "Cargo",
        "Mês",
        "Ausências",
        "Qtd. Lançamentos",
        "Elegível",
        "Assiduidade",
        "Bônus Final"
    ])

    for item in fechamento:
        ws_resumo.append([
            item["funcionario_id"],
            item["funcionario"],
            item["cargo"],
            item["mes"],
            item["ausencias"],
            item["quantidade_lancamentos"],
            "Sim" if item["elegivel"] else "Não",
            item["assiduidade"],
            item["bonus_final"]
        ])

    estilizar_header(ws_resumo)

    # Lançamentos Semanais
    ws_lanc = wb.create_sheet("Lançamentos Semanais")
    ws_lanc.append([
        "ID",
        "Funcionário ID",
        "Semana",
        "Pedidos Separados",
        "Pedidos Carregados",
        "Toneladas",
        "Entregas",
        "Retornos",
        "Nota",
        "Penalidade",
        "Motivo Penalidade",
        "Bônus Calculado"
    ])

    for l in lancamentos:
        ws_lanc.append([
            l.id,
            l.funcionario_id,
            l.semana,
            l.pedidos_separados,
            l.pedidos_carregados,
            l.toneladas,
            l.entregas,
            l.retornos,
            l.nota,
            "Sim" if l.penalidade else "Não",
            l.motivo_penalidade,
            l.bonus_calculado
        ])

    estilizar_header(ws_lanc)

    # Frequência Mensal
    ws_freq = wb.create_sheet("Frequência Mensal")
    ws_freq.append([
        "ID",
        "Funcionário ID",
        "Mês",
        "Ausências"
    ])

    for f in frequencias:
        ws_freq.append([
            f.id,
            f.funcionario_id,
            f.mes,
            f.ausencias
        ])

    estilizar_header(ws_freq)

    # Funcionários
    ws_func = wb.create_sheet("Funcionários")
    ws_func.append([
        "ID",
        "Nome",
        "Cargo",
        "Ativo"
    ])

    for func in funcionarios:
        ws_func.append([
            func.id,
            func.nome,
            func.cargo,
            "Sim" if func.ativo else "Não"
        ])

    estilizar_header(ws_func)

    for worksheet in wb.worksheets:
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            worksheet.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output