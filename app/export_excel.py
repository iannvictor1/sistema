from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def exportar_fechamento_excel(mes: str, fechamento, lancamentos, frequencias, funcionarios):
    wb = Workbook()

    ws = wb.active
    wb.remove(ws)

    cor_titulo = "1F1F1F"
    cor_header = "1F4E78"
    cor_header_clara = "D9EAF7"
    cor_elegivel = "E2F0D9"
    cor_bloqueado = "FCE4D6"
    cor_total = "FFF2CC"

    fonte_titulo = Font(color="FFFFFF", bold=True, size=14)
    fonte_header = Font(color="FFFFFF", bold=True)
    fonte_padrao = Font(color="000000")
    fonte_negrito = Font(color="000000", bold=True)

    fill_titulo = PatternFill("solid", fgColor=cor_titulo)
    fill_header = PatternFill("solid", fgColor=cor_header)
    fill_subheader = PatternFill("solid", fgColor=cor_header_clara)
    fill_elegivel = PatternFill("solid", fgColor=cor_elegivel)
    fill_bloqueado = PatternFill("solid", fgColor=cor_bloqueado)
    fill_total = PatternFill("solid", fgColor=cor_total)

    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center")

    borda_fina = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    def aplicar_borda_linha(ws, linha):
        for cell in ws[linha]:
            cell.border = borda_fina

    def estilizar_titulo(ws, titulo):
        ws.merge_cells("A1:M1")
        ws["A1"] = titulo
        ws["A1"].font = fonte_titulo
        ws["A1"].fill = fill_titulo
        ws["A1"].alignment = alinhamento_centro

    def estilizar_header(ws, linha):
        for cell in ws[linha]:
            cell.fill = fill_header
            cell.font = fonte_header
            cell.alignment = alinhamento_centro
            cell.border = borda_fina

    def formatar_moeda(cell):
        cell.number_format = 'R$ #,##0.00'

    def auto_ajustar_colunas(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                valor = "" if cell.value is None else str(cell.value)
                if len(valor) > max_length:
                    max_length = len(valor)
            ws.column_dimensions[col_letter].width = max_length + 3

    # =========================
    # Aba 1 - Resumo Fechamento
    # =========================
    ws_resumo = wb.create_sheet("Resumo Fechamento")

    estilizar_titulo(ws_resumo, "Resumo de Fechamento da Bonificação")
    ws_resumo["A2"] = "Mês de referência:"
    ws_resumo["B2"] = mes
    ws_resumo["A2"].font = fonte_negrito
    ws_resumo["A2"].fill = fill_subheader
    ws_resumo["A2"].border = borda_fina
    ws_resumo["B2"].border = borda_fina

    headers_resumo = [
        "Funcionário ID",
        "Funcionário",
        "Cargo",
        "Mês",
        "Ausências",
        "Qtd. Lançamentos",
        "Elegível",
        "Assiduidade",
        "Nota Atual",
        "Bonus Bruto",
        "Desconto",
        "Motivo Desconto",
        "Bônus Final",
    ]

    linha_header_resumo = 4
    for i, valor in enumerate(headers_resumo, start=1):
        ws_resumo.cell(row=linha_header_resumo, column=i, value=valor)

    estilizar_header(ws_resumo, linha_header_resumo)

    linha_atual = linha_header_resumo + 1
    total_bonus = 0.0
    total_assiduidade = 0.0
    total_elegiveis = 0
    total_bloqueados = 0

    for item in fechamento:
        bonus_com_assiduidade = float(item.get("bonus_bruto", item["bonus_final"]) or 0)
        assiduidade = float(item.get("assiduidade", 0) or 0)
        bonus_lancamentos = round(max(0.0, bonus_com_assiduidade - assiduidade), 2)
        elegivel = "Sim" if item["elegivel"] else "Não"

        ws_resumo.cell(linha_atual, 1, item["funcionario_id"])
        ws_resumo.cell(linha_atual, 2, item["funcionario"])
        ws_resumo.cell(linha_atual, 3, item["cargo"])
        ws_resumo.cell(linha_atual, 4, item["mes"])
        ws_resumo.cell(linha_atual, 5, item["ausencias"])
        ws_resumo.cell(linha_atual, 6, item["quantidade_lancamentos"])
        ws_resumo.cell(linha_atual, 7, elegivel)
        ws_resumo.cell(linha_atual, 8, item["assiduidade"])
        ws_resumo.cell(linha_atual, 9, item.get("nota_atual") or "-")
        ws_resumo.cell(linha_atual, 10, bonus_lancamentos)
        ws_resumo.cell(linha_atual, 11, item.get("desconto", 0))
        ws_resumo.cell(linha_atual, 12, item.get("motivo_desconto") or "-")
        ws_resumo.cell(linha_atual, 13, item["bonus_final"])

        for col in range(1, 14):
            ws_resumo.cell(linha_atual, col).border = borda_fina
            ws_resumo.cell(linha_atual, col).alignment = alinhamento_centro if col not in {2, 3, 12} else alinhamento_esquerda
            ws_resumo.cell(linha_atual, col).font = fonte_padrao

        formatar_moeda(ws_resumo.cell(linha_atual, 8))
        formatar_moeda(ws_resumo.cell(linha_atual, 10))
        formatar_moeda(ws_resumo.cell(linha_atual, 11))
        formatar_moeda(ws_resumo.cell(linha_atual, 13))

        if item["elegivel"]:
            for col in range(1, 14):
                ws_resumo.cell(linha_atual, col).fill = fill_elegivel
            total_elegiveis += 1
        else:
            for col in range(1, 14):
                ws_resumo.cell(linha_atual, col).fill = fill_bloqueado
            total_bloqueados += 1

        total_assiduidade += float(item["assiduidade"])
        total_bonus += float(item["bonus_final"])
        linha_atual += 1

    # Totais
    ws_resumo.cell(linha_atual + 1, 1, "Totais")
    ws_resumo.cell(linha_atual + 1, 5, "Elegíveis")
    ws_resumo.cell(linha_atual + 1, 6, total_elegiveis)
    ws_resumo.cell(linha_atual + 1, 7, "Bloqueados")
    ws_resumo.cell(linha_atual + 1, 8, total_bloqueados)

    ws_resumo.cell(linha_atual + 2, 7, "Total Assiduidade")
    ws_resumo.cell(linha_atual + 2, 8, total_assiduidade)

    ws_resumo.cell(linha_atual + 3, 7, "Total Bônus")
    ws_resumo.cell(linha_atual + 3, 8, total_bonus)

    for r in [linha_atual + 1, linha_atual + 2, linha_atual + 3]:
        for c in range(1, 14):
            ws_resumo.cell(r, c).border = borda_fina
            ws_resumo.cell(r, c).font = fonte_negrito
            ws_resumo.cell(r, c).fill = fill_total

    formatar_moeda(ws_resumo.cell(linha_atual + 2, 8))
    formatar_moeda(ws_resumo.cell(linha_atual + 3, 8))

    ws_resumo.freeze_panes = "A5"
    ws_resumo.auto_filter.ref = f"A4:M{max(linha_atual - 1, 4)}"
    
    funcionarios_dict = {f.id: f.nome for f in funcionarios}

    funcionarios_dict = {f.id: f.nome for f in funcionarios}

    # =========================
    # Aba 2 - Lançamentos Semanais
    # =========================
    ws_lanc = wb.create_sheet("Lançamentos Semanais")
    ws_lanc.merge_cells("A1:Q1")
    ws_lanc["A1"] = "Detalhamento dos Lançamentos Semanais"
    ws_lanc["A1"].font = fonte_titulo
    ws_lanc["A1"].fill = fill_titulo
    ws_lanc["A1"].alignment = alinhamento_centro

    headers_lanc = [
        "ID",
        "Funcionário ID",
        "Funcionário",
        "Semana",
        "Tipo",
        "Data do Lançamento",
        "Data Escolhida",
        "Usuário",
        "Pedidos Separados",
        "Pedidos Carregados",
        "Toneladas",
        "Entregas",
        "Retornos",
        "Nota",
        "Penalidade",
        "Motivo Penalidade",
        "Bônus Calculado",
    ]

    for i, valor in enumerate(headers_lanc, start=1):
        ws_lanc.cell(row=3, column=i, value=valor)

    estilizar_header(ws_lanc, 3)

    linha = 4
    for l in lancamentos:
        tipo = getattr(l, "tipo_lancamento", "semanal") or "semanal"

        if tipo == "diario":
            continue

        ws_lanc.cell(linha, 1, l.id)
        ws_lanc.cell(linha, 2, l.funcionario_id)
        ws_lanc.cell(linha, 3, funcionarios_dict.get(l.funcionario_id, "-"))
        ws_lanc.cell(linha, 4, l.semana)
        ws_lanc.cell(linha, 5, tipo)
        ws_lanc.cell(linha, 6, getattr(l, "data_registro", None))
        ws_lanc.cell(linha, 7, getattr(l, "data_lancamento", None))
        ws_lanc.cell(linha, 8, getattr(l, "usuario_lancamento", None) or "-")
        ws_lanc.cell(linha, 9, l.pedidos_separados)
        ws_lanc.cell(linha, 10, l.pedidos_carregados)
        ws_lanc.cell(linha, 11, l.toneladas)
        ws_lanc.cell(linha, 12, l.entregas)
        ws_lanc.cell(linha, 13, l.retornos)
        ws_lanc.cell(linha, 14, l.nota)
        ws_lanc.cell(linha, 15, "Sim" if l.penalidade else "Não")
        ws_lanc.cell(linha, 16, l.motivo_penalidade if l.motivo_penalidade else "-")
        ws_lanc.cell(linha, 17, l.bonus_calculado)

        for col in range(1, 18):
            ws_lanc.cell(linha, col).border = borda_fina

        formatar_moeda(ws_lanc.cell(linha, 17))
        linha += 1

    ws_lanc.freeze_panes = "A4"
    ws_lanc.auto_filter.ref = f"A3:Q{max(linha - 1, 3)}"

    # =========================
    # Aba 3 - Lançamentos Diários
    # =========================
    ws_diario = wb.create_sheet("Lançamentos Diários")
    ws_diario.merge_cells("A1:Q1")
    ws_diario["A1"] = "Detalhamento dos Lançamentos Diários"
    ws_diario["A1"].font = fonte_titulo
    ws_diario["A1"].fill = fill_titulo
    ws_diario["A1"].alignment = alinhamento_centro

    for i, valor in enumerate(headers_lanc, start=1):
        ws_diario.cell(row=3, column=i, value=valor)

    estilizar_header(ws_diario, 3)

    linha = 4
    for l in lancamentos:
        tipo = getattr(l, "tipo_lancamento", "semanal") or "semanal"

        if tipo != "diario":
            continue

        ws_diario.cell(linha, 1, l.id)
        ws_diario.cell(linha, 2, l.funcionario_id)
        ws_diario.cell(linha, 3, funcionarios_dict.get(l.funcionario_id, "-"))
        ws_diario.cell(linha, 4, l.semana)
        ws_diario.cell(linha, 5, tipo)
        ws_diario.cell(linha, 6, getattr(l, "data_registro", None))
        ws_diario.cell(linha, 7, getattr(l, "data_lancamento", None))
        ws_diario.cell(linha, 8, getattr(l, "usuario_lancamento", None) or "-")
        ws_diario.cell(linha, 9, l.pedidos_separados)
        ws_diario.cell(linha, 10, l.pedidos_carregados)
        ws_diario.cell(linha, 11, l.toneladas)
        ws_diario.cell(linha, 12, l.entregas)
        ws_diario.cell(linha, 13, l.retornos)
        ws_diario.cell(linha, 14, l.nota)
        ws_diario.cell(linha, 15, "Sim" if l.penalidade else "Não")
        ws_diario.cell(linha, 16, l.motivo_penalidade if l.motivo_penalidade else "-")
        ws_diario.cell(linha, 17, l.bonus_calculado)

        for col in range(1, 18):
            ws_diario.cell(linha, col).border = borda_fina

        formatar_moeda(ws_diario.cell(linha, 17))
        linha += 1

    ws_diario.freeze_panes = "A4"
    ws_diario.auto_filter.ref = f"A3:Q{max(linha - 1, 3)}"

    # =========================
    # Aba 4 - Frequência Mensal (RESTAURADA)
    # =========================
    ws_freq = wb.create_sheet("Frequência Mensal")
    ws_freq.merge_cells("A1:H1")
    ws_freq["A1"] = "Controle de Frequência Mensal"
    ws_freq["A1"].font = fonte_titulo
    ws_freq["A1"].fill = fill_titulo
    ws_freq["A1"].alignment = alinhamento_centro

    headers_freq = [
    "ID",
    "Funcionário ID",
    "Funcionário",
    "Mês",
    "Ausências",
    "Dia da Falta",
    "Tipo de Falta",
    "Status do Mês"
]

    for i, valor in enumerate(headers_freq, start=1):
        ws_freq.cell(row=3, column=i, value=valor)

    estilizar_header(ws_freq, 3)

    linha = 4
    for f in frequencias:
        ws_freq.cell(linha, 1, f.id)
        ws_freq.cell(linha, 2, f.funcionario_id)
        ws_freq.cell(linha, 3, funcionarios_dict.get(f.funcionario_id, "-"))
        ws_freq.cell(linha, 4, f.mes)
        ws_freq.cell(linha, 5, f.ausencias)
        ws_freq.cell(linha, 6, getattr(f, "data_falta", None))
        ws_freq.cell(
            linha,
            8,
            getattr(f, "status_mes", "Normal")
        )
        ws_freq.cell(linha, 7, getattr(f, "tipo_falta", None) or "-")

        for col in range(1, 9):
            ws_freq.cell(linha, col).border = borda_fina

        linha += 1

    ws_freq.freeze_panes = "A4"
    ws_freq.auto_filter.ref = f"A3:H{max(linha - 1, 3)}"

    ws_func = wb.create_sheet("Funcionários")
    ws_func.merge_cells("A1:E1")
    ws_func["A1"] = "Base de Funcionários"
    ws_func["A1"].font = fonte_titulo
    ws_func["A1"].fill = fill_titulo
    ws_func["A1"].alignment = alinhamento_centro

    headers_func = ["ID", "Nome", "Cargo", "Turno", "Ativo"]

    for i, valor in enumerate(headers_func, start=1):
        ws_func.cell(row=3, column=i, value=valor)

    estilizar_header(ws_func, 3)

    linha = 4
    for func in funcionarios:
        ws_func.cell(linha, 1, func.id)
        ws_func.cell(linha, 2, func.nome)
        ws_func.cell(linha, 3, func.cargo)
        ws_func.cell(linha, 4, getattr(func, "turno", "Não informado"))
        ws_func.cell(linha, 5, "Sim" if func.ativo else "Não")

        for col in range(1, 6):
            ws_func.cell(linha, col).border = borda_fina

        linha += 1

    ws_func.freeze_panes = "A4"
    ws_func.auto_filter.ref = f"A3:E{max(linha - 1, 3)}"

    for worksheet in wb.worksheets:
        auto_ajustar_colunas(worksheet)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
